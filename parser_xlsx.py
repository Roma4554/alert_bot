import logging
import os.path as path

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from datetime import datetime

from classes.Notification import Notification
from db import add_info_to_notification, get_notifications_generator


def write_from_xlsx_to_db(file_name: str) -> str:
    """
    Функция для чтения exel файла и записи данных в БД
    :param file_name: file.xlsx
    """
    notifications_list = list()

    pth = path.join(path.curdir, 'documents', file_name)

    book = load_workbook(filename=pth, data_only=True)
    sheet = book.active
    total_note = 0

    for i_row, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
        if row.count(None) == 3:
            continue

        total_note += 1
        if None not in row:
            employee_id, data, text = row
            if isinstance(data, datetime):
                data = str(data.date()).replace('-', '.')
            try:
                note = Notification(employee_id, data, text)
                notifications_list.append(tuple(note))
            except (TypeError, ValueError) as ex:
                logging.error(f'Ошибка записи данных в строке {i_row}: {ex}.\n{row}')
        else:
            logging.error(f'Ошибка записи данных в строке {i_row}: Указаны не все данные.\n{row}')

    add_info_to_notification(notifications_list)
    status = f'Запись данных завершена. В базу добавлено {len(notifications_list)} из {total_note} оповещений.'
    logging.info(status)
    return status


def write_from_db_to_xlsx() -> None:
    """
    Функция для записи информации из БД в .xlsx файл
    """

    book = Workbook()
    sheet = book.active

    # Создание шапки таблицы
    sheet['A1'].value = 'Табельный номер'
    sheet['B1'].value = 'Дата'
    sheet['C1'].value = 'Оповещение'

    cell_border = Side(style='thin', color="000000")

    for cell in sheet['A1:C1'][0]:
        cell.font = Font(name='Times New Roman', size=13, bold=True)
        cell.border = Border(left=cell_border, right=cell_border, top=cell_border, bottom=cell_border)
        cell.alignment = Alignment(horizontal='center')

    sheet.auto_filter.ref = 'A:C'

    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 13
    sheet.column_dimensions['C'].width = 80

    # Создание основного тела таблицы
    i_row = 2
    for data in get_notifications_generator():
        data = tuple(data)
        row = sheet[f'A{i_row}:C{i_row}'][0]
        cell_value = zip(row, data)
        for cell, value in cell_value:
            cell.value = value
            cell.font = Font(name='Times New Roman', size=11)
            cell.border = Border(left=cell_border, right=cell_border, top=cell_border, bottom=cell_border)
            if cell.column in (1, 2):
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(horizontal='left', wrapText=True)
        i_row += 1

    pth = path.join(path.curdir, 'documents', 'DataBase.xlsx')
    book.save(pth)
